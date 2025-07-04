import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font
from io import BytesIO
from datetime import datetime

def process_excel(uploaded_file):
    try:
        xls = pd.read_excel(uploaded_file, sheet_name=None)
    except Exception as e:
        st.error(f"讀取 Excel 檔案錯誤: {e}")
        return None

    df_fee = xls.get('費用編號表')
    if df_fee is None:
        st.error("找不到『費用編號表』分頁。")
        return None
    required_fee_columns = ['費用編號', '所屬', '項目']
    for col in required_fee_columns:
        if col not in df_fee.columns:
            st.error(f"『費用編號表』中缺少必需欄位：{col}")
            return None
    fee_mapping = df_fee.set_index('費用編號')[['所屬', '項目']].to_dict('index')

    df_customer = xls.get('客戶列表')
    if df_customer is None:
        st.error("找不到『客戶列表』分頁。")
        return None
    if '客戶名稱' not in df_customer.columns:
        st.error("『客戶列表』中缺少必需欄位：客戶名稱")
        return None

    if '客戶編號' in df_customer.columns:
        pos = df_customer.columns.get_loc('客戶編號')
        df_customer.insert(pos, '請檢查', '')
    else:
        df_customer.insert(0, '請檢查', '')
    df_customer['分倉應收賬款（未稅）'] = 0

    processed_sheets = set()

    for idx, cust_row in df_customer.iterrows():
        cust_raw = cust_row['客戶名稱']
        if pd.isna(cust_raw) or not str(cust_raw).strip():
            df_customer.loc[idx, '請檢查'] = ""
            continue

        cust_name = str(cust_raw).strip()
        cust_code = cust_name[:4]

        target_sheet_name = None
        for sheet_name in xls.keys():
            if sheet_name in ['費用編號表', '客戶列表', '營收統計']:
                continue
            if sheet_name.startswith(cust_code):
                target_sheet_name = sheet_name
                break

        if target_sheet_name is None:
            df_customer.loc[idx, '請檢查'] = "***"
            continue

        if target_sheet_name in processed_sheets:
            continue
        processed_sheets.add(target_sheet_name)

        try:
            df_cust = xls[target_sheet_name].copy()
        except Exception as e:
            st.error(f"讀取客戶分頁 {target_sheet_name} 失敗: {e}")
            continue

        if df_cust.empty:
            df_customer.loc[idx, '請檢查'] = "***"
            continue

        required_customer_columns = ['費用編號', '費用', '總計']
        missing_cols = [col for col in required_customer_columns if col not in df_cust.columns]
        if missing_cols:
            st.error(f"客戶分頁 {target_sheet_name} 缺少必需欄位：{', '.join(missing_cols)}")
            continue

        df_cust['所屬'] = df_cust['費用編號'].map(lambda x: fee_mapping[x]['所屬'] if x in fee_mapping else "")
        df_cust['項目名'] = df_cust['費用編號'].map(lambda x: fee_mapping[x]['項目'] if x in fee_mapping else "")
        fee_index = df_cust.columns.get_loc('費用編號')
        cols = list(df_cust.columns)
        if '所屬' in cols: cols.remove('所屬')
        if '項目名' in cols: cols.remove('項目名')
        new_cols = cols[:fee_index+1] + ['所屬', '項目名'] + cols[fee_index+1:]
        df_cust = df_cust[new_cols]

        total = 0
        for _, fee_row in df_cust.iterrows():
            fee_code = fee_row['費用編號']
            amount = pd.to_numeric(fee_row['總計'], errors='coerce')
            if pd.notna(amount) and fee_code in fee_mapping and fee_mapping[fee_code]['所屬'] == 2:
                total += amount
        df_customer.loc[idx, '分倉應收賬款（未稅）'] = total

        total_sum = pd.to_numeric(df_cust['總計'], errors='coerce').sum()
        subtract_sum = pd.to_numeric(df_cust.loc[df_cust['費用編號'] == 921, '總計'], errors='coerce').sum()
        final_sum = round((total_sum - subtract_sum) * 1.05)
        branch_revenue = round(total * 1.05)
        untaxed_revenue = round(total_sum - subtract_sum)

        df_cust["佔總營收%"] = df_cust.apply(
            lambda row: f"{round((pd.to_numeric(row['總計'], errors='coerce') * 1.05) / final_sum * 100, 1)}%" if final_sum != 0 else "0%", axis=1
        )
        df_cust["佔分倉營收%"] = df_cust.apply(
            lambda row: f"{round((pd.to_numeric(row['總計'], errors='coerce') * 1.05) / branch_revenue * 100, 1)}%" if branch_revenue != 0 else "0%", axis=1
        )
        if "備註" in df_cust.columns:
            pos = df_cust.columns.get_loc("備註")
            cols = list(df_cust.columns)
            cols.remove("佔總營收%")
            cols.remove("佔分倉營收%")
            new_cols = cols[:pos+1] + ["佔總營收%", "佔分倉營收%"] + cols[pos+1:]
            df_cust = df_cust[new_cols]

        new_row_untaxed = {col: "" for col in df_cust.columns}
        new_row_untaxed['費用'] = "總營收（不含稅)"
        new_row_untaxed['總計'] = untaxed_revenue

        new_row_total = {col: "" for col in df_cust.columns}
        new_row_total['費用'] = "總營收"
        new_row_total['總計'] = final_sum

        new_row_sub = {col: "" for col in df_cust.columns}
        new_row_sub['費用'] = "分倉營收"
        new_row_sub['總計'] = branch_revenue

        summary_rows = pd.DataFrame([new_row_untaxed, new_row_total, new_row_sub])
        df_cust = pd.concat([df_cust, summary_rows], ignore_index=True)

        xls[target_sheet_name] = df_cust

    df_customer['分倉應收賬款（含稅）'] = (df_customer['分倉應收賬款（未稅）'] * 1.05).round().astype(int)
    xls['客戶列表'] = df_customer

    customer_summary_list = []
    for sheet_name in processed_sheets:
        df_sheet = xls[sheet_name]
        if '費用' not in df_sheet.columns:
            continue
        row_untaxed = df_sheet[df_sheet['費用'] == "總營收（不含稅)"]
        row_total = df_sheet[df_sheet['費用'] == "總營收"]
        row_branch = df_sheet[df_sheet['費用'] == "分倉營收"]
        untaxed_val = row_untaxed.iloc[0]['總計'] if not row_untaxed.empty else 0
        total_val = row_total.iloc[0]['總計'] if not row_total.empty else 0
        branch_val = row_branch.iloc[0]['總計'] if not row_branch.empty else 0
        customer_summary_list.append({
            "客戶": sheet_name,
            "總營收（不含稅)": untaxed_val,
            "總營收": total_val,
            "分倉營收": branch_val
        })
    df_customer_summary = pd.DataFrame(customer_summary_list)
    global_untaxed = df_customer_summary["總營收（不含稅)"].sum()
    global_total = df_customer_summary["總營收"].sum()
    global_branch = df_customer_summary["分倉營收"].sum()
    global_row = {"客戶": "全局總計", "總營收（不含稅)": global_untaxed, "總營收": global_total, "分倉營收": global_branch}
    df_customer_summary = pd.concat([df_customer_summary, pd.DataFrame([global_row])], ignore_index=True)
    xls["營收統計"] = df_customer_summary

    output_buffer = BytesIO()
    with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
        for sheet_name, df in xls.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    output_buffer.seek(0)

    wb = openpyxl.load_workbook(output_buffer)
    ws_customer = wb["客戶列表"]
    header = [cell.value for cell in ws_customer[1]]
    check_idx = header.index("請檢查") + 1
    for row in ws_customer.iter_rows(min_row=2):
        if row[check_idx-1].value == "***":
            for cell in row:
                cell.font = Font(color="FF0000")

    sheet_titles = wb.sheetnames
    if "客戶列表" in sheet_titles and "營收統計" in sheet_titles:
        customer_index = sheet_titles.index("客戶列表")
        ws_summary = wb["營收統計"]
        wb._sheets.remove(ws_summary)
        wb._sheets.insert(customer_index + 1, ws_summary)

    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    return final_output

st.title("JENJAN 對賬小幫手V1 - 賬單生成")
st.write("請上傳您的 Excel 賬單（檔名不限），點擊【生成新賬單】後即可下載新賬單。")

uploaded_file = st.file_uploader("選擇 Excel 文件", type=["xlsx"], key="excel_uploader")
if uploaded_file is not None:
    if st.button("生成新賬單"):
        result = process_excel(uploaded_file)
        if result is not None:
            input_filename = uploaded_file.name
            if input_filename.lower().endswith('.xlsx'):
                input_filename = input_filename[:-5]
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"對賬後_{input_filename}_{timestamp}.xlsx"
            st.download_button(
                label="下載生成的新賬單",
                data=result,
                file_name=output_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
